import re
import requests
import zipfile
from selenium import webdriver
from selenium_stealth import stealth
from selenium.webdriver.common.by import By
from time import sleep
import numpy as np
import datetime
from random import randint
from random import uniform
import os
from selenium.webdriver.chrome.service import Service
import io
from fake_headers import Headers
from urllib.parse import unquote
import psutil
import time
from openpyxl import load_workbook
from openpyxl import Workbook
from PIL import Image

PROXY_HOST = ''
PROXY_PORT = ''
PROXY_USER = ''
PROXY_PASS = ''



manifest_json = """
{
    "version": "1.0.0",
    "manifest_version": 2,
    "name": "Chrome Proxy",
    "permissions": [
        "proxy",
        "tabs",
        "unlimitedStorage",
        "storage",
        "<all_urls>",
        "webRequest",
        "webRequestBlocking"
    ],
    "background": {
        "scripts": ["background.js"]
    },
    "minimum_chrome_version": "76.0.0"
}
"""

background_js = """
let config = {
        mode: "fixed_servers",
        rules: {
        singleProxy: {
            scheme: "http",
            host: "%s",
            port: parseInt(%s)
        },
        bypassList: ["localhost"]
        }
    };
chrome.proxy.settings.set({value: config, scope: "regular"}, function() {});
function callbackFn(details){
    return {
        authCredentials: {
            username: "%s",
            password: "%s"
        }
    };
}
chrome.webRequest.onAuthRequired.addListener(
            callbackFn,
            {urls: ["<all_urls>"]},
            ['blocking']
);
""" % (PROXY_HOST,PROXY_PORT,PROXY_USER,PROXY_PASS)


def get_chromedriver(use_proxy=False):
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("start-maximized")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("--disable-notifications")
    print('Агент нужен? \n 1 - Yes \n 2 - No')
    agent = int(input('Выбор = '))
    if agent == 1:
        chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36")

    if use_proxy:
        plugin_file = 'proxy_auth_plugin.zip'
        
        with zipfile.ZipFile(plugin_file, 'w') as zp:
            zp.writestr('manifest.json', manifest_json)
            zp.writestr('background.js', background_js)
            
        chrome_options.add_extension(plugin_file)
        
    #if user_agent:
        #chrome_options.add_argument(f'--user-agent={user_agent}')
        
    #s = Service(
    #    executable_path='.'
    #)
    driver = webdriver.Chrome(
        #service=s,
        options=chrome_options
    )
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
    'source': '''
        delete window.cdc_adoQpoasnfa76pfcZLmcfl_Array;
        delete window.cdc_adoQpoasnfa76pfcZLmcfl_Object;
        delete window.cdc_adoQpoasnfa76pfcZLmcfl_Promise;
        delete window.cdc_adoQpoasnfa76pfcZLmcfl_Proxy;  
        delete window.cdc_adoQpoasnfa76pfcZLmcfl_Symbol;      
    '''
    })
    stealth(driver,
    languages=["en-US","en"],
    vendor="Yandex llc.",
    platform="Win32",
    webgl_vendor="Intel Inc.",
    renderer="Intel Iris OpenGL Engine",
    fix_hairline=True,)
    return driver

headers = {
"accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7", "User-Agent": "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/527.36 (KHTML, like Gecko) Chrome/114.0.1985.67 Safari/527.36"
}

def exel_exist_check(choose):
    if int(choose) == 2:
        path = 'AvitoParser.xlsx'
        try:
            excel = load_workbook(path)
            print('Файл AvitoParser.xlsx уже существует !')
            return('-') # Чтобы не сбивать логику return'ov
        except:
            print('Эксель файла AvitoParser.xlsx для записи не обнаружено, идет создание нового файла')
            excel = Workbook()
            ws = excel.active
            ws.title = "Test"
            excel.save('AvitoParser.xlsx')
            print('Создание AvitoParser.xlsx прошло успешно')
            return('-') #Возвращает '-' в знак того, что файлик только что был создан
    else:
        path = 'AvitoAnalyzer.xlsx'
        try:
            excel = load_workbook(path)
            print('Файл AvitoAnalyzer.xlsx уже существует !')
            sheet = excel['data']
            columns = sheet.max_column
            rows = sheet.max_row
            print(rows)
            print(columns)
            print('Значение первой строки')
            first_string = []
            final_string = []
            exist_marks = []
            exist_models = []
            exist_generations = []

            for i in range(1, columns + 1): 
                first_string.append(sheet.cell(row=2, column=i))
                final_string.append(sheet.cell(row=rows, column=i))  
            print(*[i.value for i in first_string])
            print('Значение последней строки')
            print(*[i.value for i in final_string])

            for i in range(1, rows + 1): 
                exist_marks.append(sheet.cell(row=i, column=1))
                exist_models.append(sheet.cell(row=i, column=2)) 
                exist_generations.append(sheet.cell(row=i, column=3))
            exist_marks = [i.value for i in exist_marks]
            exist_models = [i.value for i in exist_models]
            exist_generations = [i.value for i in exist_generations]

            return(rows,exist_marks,exist_models,exist_generations) #Возвращает количество строк (записей в файлике)
        except:
            print('Эксель файла AvitoAnalyze.xlsx для записи не обнаружено, идет создание нового файла')
            excel = Workbook()
            ws = excel.active
            ws.title = "data"
            excel.save('AvitoAnalyzer.xlsx')
            print('Создание AvitoParser.xlsx прошло успешно')
            excel = load_workbook(path)
            sheet = excel['data']
            sheet['A1'] = 'Марка'
            sheet['B1'] = 'Модель'
            sheet['C1'] = 'Поколение'
            sheet['D1'] = 'Цена в мск'
            sheet['E1'] = 'Кол-во'
            sheet['F1'] = 'Питер'
            sheet['G1'] = 'Кол-во'
            sheet['H1'] = 'Астрахань'
            sheet['I1'] = 'Кол-во'
            sheet['J1'] = 'Армавир'
            sheet['K1'] = 'Кол-во'
            sheet['L1'] = 'Краснодар'
            sheet['M1'] = 'Кол-во'
            sheet['N1'] = 'Махачкала'
            sheet['O1'] = 'Кол-во'
            excel.save('AvitoAnalyzer.xlsx')
            return('-') #Возвращает '-' в знак того, что файлик только что был создан
def Output(avitolink, avitotitle, avitoprice, data, text, gen):
    path = 'AvitoParser.xlsx'
    excel = load_workbook(path)
    try:
        sheet = excel[f'{text} {gen}']
    except:
        new_sheet = excel.create_sheet(f'{text} {gen}')
    excel.save(path)
    sheet = excel[f'{text} {gen}']
    for i in range(len(avitolink)):
        sheet['A{}'.format(i+1)] = '{}'.format(avitolink[i])
        sheet['B{}'.format(i+1)] = '{}'.format(avitotitle[i])
        sheet['C{}'.format(i+1)] = '{}'.format(avitoprice[i][0])
        sheet['D{}'.format(i+1)] = '{}'.format(data[i])
    excel.save(path)
    excel.close


def AvitoParser(text,driver,min_price,need_price_filter,max_price,gen):
    avitoprice = []
    avitotitle = []
    avitolink = []
    data = []
    sleep(uniform(1,4))
    confirm = driver.find_element(By.CLASS_NAME, 'desktop-9uhrzn')
    confirm.click() #Для снятия бага word word word word
    sleep(uniform(3,5))
    gen_index = '-'
    if gen != '-': #Если мы ввели поколение
        exist_generation = []
        generation_web_element = []
        for v in range(999):
            try:
                exist_generation.append(driver.find_element(By.XPATH, '//*[@id="app"]/div/div[8]/div/div[2]/div[3]/div[3]/div[1]/div[2]/div[{}]/a'.format(v+1)).text) #Поиск всех поколений введеной модели
                generation_web_element.append(driver.find_element(By.XPATH, '//*[@id="app"]/div/div[8]/div/div[2]/div[3]/div[3]/div[1]/div[2]/div[{}]/a'.format(v+1))) #Поиск всех веб-элементов поколений введеной модели
            except:
                break
        print(f'Существующие поколения {exist_generation}')
        if len(exist_generation) == 0: #Не нашли поколения
            print(f'Поколения у данной модели отсутствует. Производится поиск по {text}')
        else:
            for g in range(len(exist_generation)): #Если введенное поколение содержится в существующих, то все ок, иначе вывод, что его нет. Проверка осуществляется по вхождению
                if gen in exist_generation[g]:
                    gen_index = g
                    break
        if gen_index == '-':
            print(f'Введеного поколения {gen} у машины {text} не существует. Проводится поиск по запросу {text}')
        else:
            generation_web_element[gen_index].click()
    sleep(uniform(3,5))
    if need_price_filter == '1':
        min_max_filter = driver.find_element(By.CLASS_NAME,'form-mainFilters-okGYr').find_elements(By.CLASS_NAME, 'styles-module-root-iJOln.styles-module-root_dense-zQwyK.styles-module-root_compensate_bottom-DviN7')[6] 
        driver.execute_script("arguments[0].scrollIntoView(true);",min_max_filter) #Опускаемся до строки фильтра цены
        sleep(uniform(0.5,1))
        print(min_price,'\n',max_price)
        min_price_button = driver.find_element(By.CLASS_NAME,'form-mainFilters-okGYr').find_elements(By.CLASS_NAME, 'styles-module-root-iJOln.styles-module-root_dense-zQwyK.styles-module-root_compensate_bottom-DviN7')[6].find_element(By.CLASS_NAME, 'styles-module-col-_bCZg.styles-module-col_span_12-ot4_6').find_element(By.XPATH,'//*[@id="app"]/div/div[8]/div/div[2]/div[3]/div[1]/div/div[3]/div[2]/div[2]/form/div[8]/div[2]/div/div/div/label[1]').find_element(By.CLASS_NAME, 'styles-module-input-Lisnt')
        min_price_button.click()
        sleep(uniform(0.5,1))
        for i in min_price:
            min_price_button.send_keys('{}'.format(i))
            sleep(uniform(0.5,1))
        max_price_button = driver.find_element(By.CLASS_NAME,'form-mainFilters-okGYr').find_elements(By.CLASS_NAME, 'styles-module-root-iJOln.styles-module-root_dense-zQwyK.styles-module-root_compensate_bottom-DviN7')[6].find_element(By.CLASS_NAME, 'styles-module-col-_bCZg.styles-module-col_span_12-ot4_6').find_element(By.XPATH,'//*[@id="app"]/div/div[8]/div/div[2]/div[3]/div[1]/div/div[3]/div[2]/div[2]/form/div[8]/div[2]/div/div/div/label[2]').find_element(By.CLASS_NAME, 'styles-module-input-Lisnt')
        max_price_button.click()
        sleep(uniform(0.5,1))
        for i in max_price:
            max_price_button.send_keys('{}'.format(i))
            sleep(uniform(0.5,1))
        sleep(uniform(0.5,1.5))
        filters_confirm = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[8]/div/div[2]/div[3]/div[1]/div/div[3]/div[2]/div[4]/button/span/span')
        sleep(uniform(0.5,1.5))
        filters_confirm.click()
        sleep(uniform(0.5,1.5))
        confirm = driver.find_element(By.CLASS_NAME, 'desktop-9uhrzn')
        driver.execute_script("arguments[0].scrollIntoView(true);",confirm) #Поднимаемся на самый верх
    try:
        sleep(uniform(1,2))
        sort_button = driver.find_element(By.CLASS_NAME, 'index-topPanel-McfCA').find_element(By.CLASS_NAME, 'styles-module-theme-rOnN1')
        sleep(uniform(2,4))
        sort_button.click()
        sleep(uniform(2,4))
        up_price_filter = driver.find_element(By.XPATH, '/html/body/div[3]/div[2]/div/div/div/button[2]/div')
        sleep(uniform(2,4))
        up_price_filter.click()
    except:
        pass

    try:
        avitoallitems = driver.find_element(By.CLASS_NAME, 'items-items-kAJAg').find_elements(By.CLASS_NAME, 'iva-item-content-rejJg')
        avito = 1
    except:
        avito = 0
        avitolink = '-'
        avitoprice = ['-']
    if avito == 1:
        print(len(avitoallitems))
        limit = len(avitoallitems)
        if len(avitoallitems) >= 30:
            limit = 30
        for j in range(limit):
            avitotitle.append(avitoallitems[j].find_element(By.CLASS_NAME, 'iva-item-title-py3i_').text)
            trashprice = ((((avitoallitems[j].find_element(By.CLASS_NAME, 'price-price-JP7qe').text).replace(" ","")).replace("Бесплатно","0")).replace("Ценадоговорная","0")) #margin-right:4px
            avitoprice.append(re.findall('([0-9].*?)₽',trashprice))
            avitolink.append(avitoallitems[j].find_element(By.CLASS_NAME, 'styles-module-root-YeOVk.styles-module-root_noVisited-MpiGq').get_attribute('href'))
            data.append(avitoallitems[j].find_element(By.CLASS_NAME,'iva-item-dateInfoStep-_acjp').text)
            sleep(randint(4,6)/randint(2,3))

    Output(avitolink, avitotitle, avitoprice, data, text, gen)


def AvitoAnalyze(driver,counter,excel_data):
    data_checked = 0 #Флаг, необходимый для того, чтобы проверка не осуществлялась каждый раз
    ### Для адекватного представления
    confirm = driver.find_element(By.CLASS_NAME, 'desktop-9uhrzn')
    confirm.click()
    sleep(uniform(2,3))
    ###
    allbutton = driver.find_element(By.CLASS_NAME, 'popular-rubricator-button-WWqUy') #'//*[@id="app"]/div/div[7]/div/div[2]/div[3]/div[3]/div[1]/div[1]/button')
    allbutton.click()
    marks = []
    sleep(uniform(1,2))
    for i in range(999): 
        try:
            markcount = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[8]/div/div[2]/div[3]/div[3]/div[1]/div[2]/div[{}]/span'.format(i+1)).text
            if int(markcount.replace(" ","")) >= 300:
                marks.append(driver.find_element(By.XPATH, '//*[@id="app"]/div/div[8]/div/div[2]/div[3]/div[3]/div[1]/div[2]/div[{}]/a'.format(i+1)).text)
            else:
                continue
        except:
            break
    print(f'Марки - {marks}')
    flag = 0 #Флаг, который служит стопом, при найденной модели
    for i in range(len(marks)):
        if excel_data != '-' and data_checked == 0 and flag == 0:
            print(f'Текущая марка {marks[i]}, необходимая марка {excel_data[1][-1]}')
            if marks[i] != excel_data[1][-1]: #Проверяем, чтобы марка соответствовала последней отпаршеной марке
                print('Не равны')
                continue
            else:
                flag = 1
                print(f'Марка {marks[i]} соответствует последней марке {excel_data[1][-1]}')

        driver.switch_to.window(driver.window_handles[0])
        error = 0
        request_input = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[8]/div/div[1]/div/div[3]/div[2]/div[1]/div/div/div/label[1]/input')
        request_input.click()
        sleep(uniform(0.5,1))
        request_input.send_keys('{}'.format(marks[i]))
        sleep(uniform(2,4))
        confirm = driver.find_element(By.CLASS_NAME, 'desktop-9uhrzn')
        confirm.click()
        sleep(uniform(2,4))
        try:
            allbutton = driver.find_element(By.CLASS_NAME, 'popular-rubricator-button-WWqUy')
            allbutton.click()
        except:
            pass
        models = []
        for j in range(999):
            try:
                modelcount = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[8]/div/div[2]/div[3]/div[3]/div[1]/div[2]/div[{}]/span'.format(j+1)).text
                print(modelcount)
                if int(modelcount.replace(" ","")) >= 40:
                    models.append(driver.find_element(By.XPATH, '//*[@id="app"]/div/div[8]/div/div[2]/div[3]/div[3]/div[1]/div[2]/div[{}]/a'.format(j+1)).text)
                else:
                    continue
            except:
                break
        for j in models:
            if j == marks[i]:
                print(j,'=', marks[i])
                error = 1
                break
        if error == 1:
            print('Что-то пошло не так, скип марки')
            driver.get('https://www.avito.ru/all/avtomobili')
            sleep(5)
            data_checked = 1
            continue 
        print(f'модели{models}')
        if len(models) == 0:
            sleep(uniform(2,3))
            driver.get('https://www.avito.ru/all/avtomobili')
            sleep(uniform(2,3))
            continue
        flag = 0
        for j in range(len(models)): #Поиск
            if excel_data != '-' and data_checked == 0 and flag == 0:
                if models[j] != excel_data[2][-1]: #Проверяем, чтобы модель соответствовала последней отпаршеной модели
                    continue
                else:
                    flag = 1
                    print(f'Модель {models[j]} соответствует последней модели {excel_data[2][-1]}')
            driver.switch_to.window(driver.window_handles[0])
            request_input = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[8]/div/div[1]/div/div[3]/div[2]/div[1]/div/div/div/label[1]/input')
            request_input.click()
            sleep(uniform(0.5,1))
            request_input.send_keys('{} {}'.format(marks[i],models[j]))
            sleep(uniform(2,4))
            confirm = driver.find_element(By.CLASS_NAME, 'desktop-9uhrzn')
            confirm.click()
            sleep(uniform(3,6))
            generation = [] 
            for v in range(999):
                try:
                    generation_count = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[8]/div/div[2]/div[3]/div[3]/div[1]/div[2]/div[{}]/span'.format(v+1)).text
                    print(generation_count)
                    if int(generation_count.replace(" ","")) >= 15:
                        generation.append(driver.find_element(By.XPATH, '//*[@id="app"]/div/div[8]/div/div[2]/div[3]/div[3]/div[1]/div[2]/div[{}]/a'.format(v+1)).text)
                    else:
                        continue
                except:
                    break
            print(f'Поколения : {generation}')
            if len(generation) == 0:
                sleep(uniform(2,3))
                driver.get('https://www.avito.ru/all/avtomobili')
                sleep(uniform(2,3))
                continue
            #################################
            flag = 0
            for v in range(len(generation)):
                if excel_data != '-' and data_checked == 0 and flag == 0:
                    if generation[v] != excel_data[3][-1]: #Проверяем, чтобы поколение соответствовала последнему отпаршенному поколению
                        continue
                    else:
                        print(f'Поколение {generation[v]} соответствует последнему поколению {excel_data[3][-1]}')
                if excel_data != '-' and data_checked == 0 and flag == 0:
                    if generation[v] == excel_data[3][-1]: #Скпиаем последнее записанное поколение
                        flag = 1
                        data_checked = 1
                        print(f'Пропускаем поколение {generation}')
                        continue
                print(f'Текущая рассматриваемая машина -> {marks[i]} {models[j]} {generation[v]}')
                regions = ['moskva','sankt-peterburg','astrahan', 'armavir','krasnodar','mahachkala']
                flag2 = 0
                for z in range(len(regions)):
                    region_price = region_parser(driver,regions[z],generation[v], models[j], marks[i],len(generation))
                    if z == 0:
                        moskow_price = region_price[0]
                        moskow_count = region_price[1]
                        if moskow_price == 0 or moskow_price == '0':
                            print("Цена в Москве равна 0, скип поколения")
                            flag2 = 1
                            break
                    elif z == 1:
                        sankt_peterburg_price = region_price[0]/moskow_price
                        sankt_peterburg_count = region_price[1]
                    elif z == 2:
                        astrahan = region_price[0]/moskow_price
                        astrahan_count = region_price[1]
                    elif z == 3:
                        armavir = region_price[0]/moskow_price
                        armavir_count = region_price[1]
                    elif z == 4:
                        krasnodar = region_price[0]/moskow_price
                        krasnodar_count = region_price[1]
                    elif z == 5:
                        mahachkala = region_price[0]/moskow_price
                        mahachkala_count = region_price[1]
                if flag2 == 0: ##Для скипа, когда цена в Мск = 0
                    print(moskow_price,sankt_peterburg_price,astrahan,armavir,krasnodar,mahachkala,moskow_count,sankt_peterburg_count,astrahan_count,armavir_count,krasnodar_count,mahachkala_count)

                    Analyze_record(marks[i],models[j],generation[v],moskow_price,sankt_peterburg_price,astrahan,armavir,krasnodar,mahachkala,counter,moskow_count,sankt_peterburg_count,astrahan_count,armavir_count,krasnodar_count,mahachkala_count)

                    counter += 1


def Analyze_record(marks,models,generation,moskow_price,sankt_peterburg_price,astrahan,armavir,krasnodar,mahachkala,counter,moskow_count,sankt_peterburg_count,astrahan_count,armavir_count,krasnodar_count,mahachkala_count):
    path = 'AvitoAnalyzer.xlsx'
    excel = load_workbook(path)
    sheet = excel['data']
    sheet['A{}'.format(counter)] = '{}'.format(marks)
    sheet['B{}'.format(counter)] = '{}'.format(models)
    sheet['C{}'.format(counter)] = '{}'.format(generation)
    sheet['D{}'.format(counter)] = '{}'.format(moskow_price) 
    sheet['E{}'.format(counter)] = '{}'.format(moskow_count) # moskow_count
    sheet['F{}'.format(counter)] = '{}'.format(sankt_peterburg_price) # sankt_peterburg_price
    sheet['G{}'.format(counter)] = '{}'.format(sankt_peterburg_count) # sankt_peterburg_count
    sheet['H{}'.format(counter)] = '{}'.format(astrahan) # astrahan
    sheet['I{}'.format(counter)] = '{}'.format(astrahan_count) # astrahan_count
    sheet['J{}'.format(counter)] = '{}'.format(armavir) # armavir
    sheet['K{}'.format(counter)] = '{}'.format(armavir_count) # armavir_count
    sheet['L{}'.format(counter)] = '{}'.format(krasnodar) # krasnodar
    sheet['M{}'.format(counter)] = '{}'.format(krasnodar_count) # krasnodar_count
    sheet['N{}'.format(counter)] = '{}'.format(mahachkala) # mahachkala
    sheet['O{}'.format(counter)] = '{}'.format(mahachkala_count) # mahachkala_count
    excel.save(path)
    excel.close


def region_parser(driver, regions, generation, models, marks, len_gen):
        print(regions + '\n' + models)
        link2 = 'https://www.avito.ru/{}/avtomobili?q={} {}'.format(regions,marks,models)
        try:
            driver.execute_script("window.open('{}','secondtab');".format(link2))
        except:
            driver.execute_script("window.open('{}'.format(link2),'secondtab');")
        driver.switch_to.window('secondtab')
        avitoprice = []
        if len_gen > 1:
            try:
                sleep(uniform(5,6))
                gen_input = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[8]/div/div[2]/div[3]/div[1]/div/div[3]/div[2]/div[2]/form/div[7]/div[2]/label')
                sleep(uniform(1,2))
                driver.execute_script("arguments[0].scrollIntoView(true);",gen_input)
                sleep(uniform(1,2))
                gen_input.click()
                sleep(uniform(1,2))
                gen_input_after_click = driver.find_element(By.XPATH,'//*[@id="app"]/div/div[8]/div/div[2]/div[3]/div[1]/div/div[3]/div[2]/div[2]/form/div[7]/div[2]/label/div/div/input')#'//*[@id="app"]/div/div[7]/div/div[2]/div[3]/div[1]/div/div[3]/div[2]/div[2]/form/div[6]/div[2]/label/div/div/input')
                gen_input_after_click.send_keys('{}'.format(generation))
                sleep(uniform(1,2))
                first = driver.find_element(By.CLASS_NAME, 'multiselect-image-item__photo-PnJRM')
                sleep(0.5)
                first.click()
                sleep(uniform(1,2))
                second = driver.find_element(By.CLASS_NAME, 'styles-module-root-_hgvD.styles-module-root_size_m-px1W7.styles-module-root_preset_accent-anIU_')
                sleep(1)
                second.click()
                sleep(uniform(1,2))
                confirm = driver.find_element(By.CLASS_NAME, 'desktop-9uhrzn')
                driver.execute_script("arguments[0].scrollIntoView(true);",confirm)
            except:
                print('Цена = 0')
                avg_price = 0
                return(avg_price,0)

        confirm = driver.find_element(By.CLASS_NAME, 'desktop-9uhrzn')
        confirm.click()
        sleep(uniform(1,2))
        try:
            avitoallitems = driver.find_element(By.CLASS_NAME, 'items-items-kAJAg').find_elements(By.CLASS_NAME, 'iva-item-content-rejJg')
        except:
            print('netuy')
            avg_price = 0
            return(avg_price,0)
        print(len(avitoallitems))
        if len(avitoallitems) == 0:
            print('netuy')
            avg_price = 0
            return(avg_price,0)
        for j in range(len(avitoallitems)):
            trashprice = ((((avitoallitems[j].find_element(By.CLASS_NAME, 'price-price-JP7qe').text).replace(" ","")).replace("Бесплатно","0")).replace("Ценадоговорная","0")) #margin-right:4px
            avitoprice.append(re.findall('([0-9].*?)₽',trashprice))
            sleep(randint(4,6)/randint(2,3))
        avg_price = sum([int(i[0]) for i in avitoprice])/len(avitoprice)
        return(avg_price,len(avitoallitems))

def main():
    choose = input('Анализ рынка - 1, Парсер определенного региона - 2 -> ')
    if choose == '2':
        text = input('Название машины ->')
        city = input('Название города ->')
        need_price_filter = input('Нужен фильтр по цене? Да - 1, нет - 2 -> ')
        if need_price_filter == '1':
            min_price = input('Минимальная цена -> ')
            max_price = input('Максимальная цена -> ')
        else:
            min_price = max_price = 0
        choose2 = input('Есть поколение - 1, Нет поколения- 2 -> ')
        if choose2 == '1':
            gen = input('Поколение ->')
        else:
            gen = '-'
        rows_count = exel_exist_check(choose)
        driver = get_chromedriver(use_proxy=False)
        sleep(randint(1,3))
        moscow_url = 'https://www.avito.ru/{}/avtomobili?q={}'.format(city,text)
        driver.get(moscow_url)
        AvitoParser(text,driver,min_price,need_price_filter,max_price,gen)
    else:
        excel_data = exel_exist_check(choose)
        if excel_data == '-':
            counter = 2
        else:
            counter = int(excel_data[0])
        driver = get_chromedriver(use_proxy=False)
        driver.get('https://www.avito.ru/all/avtomobili')
        sleep(uniform(2,3))
        AvitoAnalyze(driver,counter,excel_data)

        
if __name__ == "__main__":
    main()